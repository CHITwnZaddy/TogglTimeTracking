import io
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


FONT_NAME = "Actor Regular"
FONT_SIZE = 12


def _fill(hex_color: str) -> PatternFill:
    """Return a solid PatternFill for the given hex color (no # prefix)."""
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _apply_row_fill(ws, row: int, fill: PatternFill, col_start: int = 1, col_end: int = 10) -> None:
    """Paint every cell in a row with the given fill."""
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = fill


def _seconds_to_time(seconds: int) -> datetime.time:
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    # Excel time values cap at 23:59 via datetime.time; for durations > 24h use [h]:mm format
    return datetime.time(min(hours, 23), minutes)


def generate_invoice(
    entries: list,
    invoice_number: str,
    period_start: datetime.date,
    period_end: datetime.date,
    hourly_rate: float,
) -> bytes:
    """Generate an invoice .xlsx matching the STL invoice template. Returns raw bytes."""
    wb = Workbook()
    ws = wb.active
    # Match original sheet name pattern
    ws.title = f"Time entries from {period_start.strftime('%m_%d_%Y')}..."

    bold = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    regular = Font(name=FONT_NAME, size=FONT_SIZE)
    vcenter = Alignment(vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # ── Row fills ─────────────────────────────────────────────────────────────
    header_fill = _fill("C3C4EB")   # header row
    white_fill   = _fill("FFFFFF")  # odd data rows
    alt_fill     = _fill("EFF0FA")  # even data rows
    total_fill   = _fill("E1E1F5")  # total row

    # ── Row 1: INVOICE # ────────────────────────────────────────────────────
    ws["A1"] = "INVOICE #: "
    ws["A1"].font = bold
    ws["A1"].alignment = vcenter
    ws["B1"] = invoice_number
    ws["B1"].font = regular
    ws["B1"].alignment = vcenter
    ws.merge_cells("B1:I1")
    ws.row_dimensions[1].height = 26

    # ── Row 2: spacer ───────────────────────────────────────────────────────
    ws["B2"] = ""
    ws["B2"].font = regular
    ws.merge_cells("B2:I2")
    ws.row_dimensions[2].height = 26

    # ── Row 3: PERIOD ───────────────────────────────────────────────────────
    ws["A3"] = "PERIOD:"
    ws["A3"].font = bold
    ws["A3"].alignment = vcenter
    ws["B3"] = (
        f"{period_start.strftime('%m/%d/%Y')} - {period_end.strftime('%m/%d/%Y')}"
    )
    ws["B3"].font = regular
    ws["B3"].alignment = vcenter
    ws.merge_cells("B3:I3")
    ws.row_dimensions[3].height = 26

    # ── Row 5: Column headers ───────────────────────────────────────────────
    # Row 4 is blank; headers go in row 5 to match template
    headers = {
        2: "Start date",
        3: "Stop date",
        4: "Project",
        5: "Task",
        6: "Description",
        7: "Duration",
        8: "Member",
        9: "Amount",
    }
    _apply_row_fill(ws, 5, header_fill)
    for col, label in headers.items():
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = bold
        cell.alignment = center
    ws.row_dimensions[5].height = 27

    # ── Data rows ────────────────────────────────────────────────────────────
    data_start = 6
    for i, entry in enumerate(entries):
        row = data_start + i
        ws.row_dimensions[row].height = 27

        # Alternating row background: row 0 → white, row 1 → light purple, etc.
        row_fill = white_fill if i % 2 == 0 else alt_fill
        _apply_row_fill(ws, row, row_fill)

        start_val = entry["start"].replace(tzinfo=None)
        stop_val = entry["stop"].replace(tzinfo=None) if entry["stop"] else None

        cells = [
            (2, start_val, "DD MMM YY", left),
            (3, stop_val, "DD MMM YY", left),
            (4, entry["project"], "General", left),
            (5, entry["task"], "General", left),
            (6, entry["description"], "General", left),
            (7, _seconds_to_time(entry["duration_seconds"]), 'h:mm;@', left),
            (8, entry["member"], "General", left),
        ]
        for col, val, fmt, align in cells:
            c = ws.cell(row=row, column=col, value=val)
            c.font = bold if col == 2 else regular   # Bold start dates (col B)
            c.alignment = align
            c.number_format = fmt

        hours = entry["duration_seconds"] / 3600
        amount = round(hours * hourly_rate, 2)
        amt_cell = ws.cell(row=row, column=9, value=amount)
        amt_cell.font = bold   # Bold row amounts
        amt_cell.alignment = vcenter
        amt_cell.number_format = '#,##0.00\\ "USD"'

    # ── Total row ────────────────────────────────────────────────────────────
    total_row = data_start + len(entries)
    ws.row_dimensions[total_row].height = 27
    _apply_row_fill(ws, total_row, total_fill)

    f_total = ws.cell(row=total_row, column=6, value="Total")
    f_total.font = bold
    f_total.alignment = right

    # Calculate total hours in Python and round to the nearest quarter hour (0.25).
    # Avoids Excel time-fraction arithmetic which caused the ":20" display bug.
    total_seconds = sum(e["duration_seconds"] for e in entries)
    total_hours_rounded = round(total_seconds / 3600 * 4) / 4

    g_total = ws.cell(row=total_row, column=7, value=total_hours_rounded)
    g_total.font = bold
    g_total.alignment = left
    g_total.number_format = '0.00" hrs"'

    h_total = ws.cell(row=total_row, column=8, value="Total")
    h_total.font = bold
    h_total.alignment = right

    sheet_name = ws.title
    i_total = ws.cell(
        row=total_row,
        column=9,
        value=f"=SUBTOTAL(109,'{sheet_name}'!$I${data_start}:$I${total_row - 1})",
    )
    i_total.font = bold
    i_total.alignment = vcenter
    i_total.number_format = '#,##0.00\\ "USD"'

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "A": 10.83,
        "B": 18.83,
        "C": 18.83,
        "D": 18.83,
        "E": 32.83,
        "F": 52.83,
        "G": 13.83,
        "H": 26.83,
        "I": 15.83,
        "J": 10.83,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
